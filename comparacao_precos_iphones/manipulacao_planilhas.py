import openpyxl

# Função para criar o workbook e adicionar uma planilha
def create_sheet():
    book = openpyxl.Workbook()
    print(book.sheetnames)
    return book

# Função para adicionar uma nova linha (novo modelo) à planilha
def adicionar_novo_modelo(iphones_page, dados):
    iphones_page.append(dados)
    print(f"Modelo {dados[1]} adicionado com sucesso!")

# Função para atualizar preços de um modelo existente
def atualizar_preco(iphones_page, indice, loja1, loja2, loja3):
    for row in iphones_page.iter_rows(min_row=2):
        if row[0].value == indice:
            row[2].value = loja1
            row[3].value = loja2
            row[4].value = loja3
            print(f"Preços do modelo {row[1].value} atualizados.")

# Função para excluir um modelo da planilha
def excluir_modelo(iphones_page, indice):
    for row in iphones_page.iter_rows(min_row=2):
        if row[0].value == indice:
            iphones_page.delete_rows(row[0].row, 1)
            print(f"Modelo {indice} removido.")

# Função para filtrar e exibir modelos com preço até um valor máximo
def filtrar_por_preco(iphones_page, preco_maximo):
    for row in iphones_page.iter_rows(min_row=2):
        preco = float(row[2].value.replace('R$', '').replace(',', ''))
        if preco <= preco_maximo:
            print(f'{row[0].value} ~ {row[1].value} ~ {row[2].value}')

# Função para salvar o arquivo Excel
def salvar_arquivo(book, nome_arquivo):
    book.save(nome_arquivo)
    print(f"Arquivo {nome_arquivo} salvo com sucesso.")
